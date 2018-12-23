import xlrd
import xlwt
#import xlwt
import sys
from datetime import datetime
from xlutils.copy import copy
import openpyxl

#Resolving issue with xlrd import in Python 2.7
#open this link https://bootstrap.pypa.io/get-pip.py and save as get-pip.py and copy this file into C:\Python2.7\
#C:\Python2.7\python.exe get-pip.py
#After this pip is installed in your system now installing xlrd
#C:\Python2.7\python.exe -m pip install xlrd
#Open python and import xlrd
#import xlrd
#it will work.

def parse_excel(path, sheet_nr, searched_text):
    # excel filen som ett objekt
    workbook = xlrd.open_workbook(path)

    # vilket blad nr som arbetet ska göras på
    sheet = workbook.sheet_by_index(sheet_nr)

    # funktion som skriver ut lite info om filen som öppnas
    print_stats(path, workbook, sheet_nr)

    #list_data = []
    list_data = {}
    inx = 0
    for row_index in range(0, sheet.nrows):
        #row = 
        add = False
        for col_index in range(sheet.ncols):
            print(sheet.cell(row_index, col_index).value)
            if (sheet.cell(row_index, col_index).value == searched_text):
                #tmp_list.append[[row_index, col_index, sheet.cell(row_index, col_index+1).value]]
                add = True
                date = sheet.cell(row_index, col_index+1).value
                date_tuple = xlrd.xldate_as_tuple(date, workbook.datemode)
                c = datetime(*date_tuple) 
                print("text ", sheet.cell(row_index, col_index).value, " data ", sheet.cell(row_index, col_index).value)
                #list_data[inx] = {}
                list_data = {inx : {"rad":row_index, "column":col_index, "searched_text":sheet.cell(row_index, col_index).value,"value":c.strftime("%d-%m-%Y")}}
                inx = inx + 1

    print_result(list_data)

def print_result(list_data):
    print("\n-------{0:-<50}".format(""))
    print("Results:")
    for key, value in list_data.items():
        for k in value.keys():
            print('{:>15}{:>25}'.format(k, value.get(k)))

def print_stats(file, workbook, sheet_nr):
    print('-Excel file that is opened: {0}\n-Sheet that data will be extracted from: {1}\n-Total number of sheets in the excel file: {2}\n-Name of all the sheets:\n{3}'.format(file, sheet_nr, workbook.nsheets,
        "\n".join('{}: {}'.format(*k) for k in enumerate(workbook.sheet_names()))) )

def create_new_xls_file():
    workbook = xlwt.Workbook()
    workbook.save('my_file.xls')
    #sheet = workbook.add_sheet('Sheet_1')

def fetch_raw_data(path, sheet_nr, col_name):
    # excel filen som ett objekt
    workbook = xlrd.open_workbook(path)

    # vilket blad nr som arbetet ska göras på
    sheet = workbook.sheet_by_index(sheet_nr)

    # funktion som skriver ut lite info om filen som öppnas
    print_stats(path, workbook, sheet_nr)

    list_data = []

    for col_index in range(sheet.ncols):
        col = sheet.cell_value(0,col_index)
        if (col_name == col):
            print(sheet.cell_value(0,col_index))
            data = []
            for row_index in range(1, sheet.nrows):
                data.append(sheet.cell(row_index, col_index).value)
            list_data.append([sheet.cell(0, col_index).value, data])
    print(list_data)
    return list_data

def insert_data_to_plot_file(path, sheet_nr, raw_data):
    print("\ninserting data in plot file")
    # excel filen som ett objekt
    workbook = openpyxl.load_workbook(path, data_only=True)
    # vilket blad nr som arbetet ska göras på
    sheet = workbook.get_sheet_by_name("Blad1")

    # funktion som skriver ut lite info om filen som öppnas
    #print_stats(path, workbook, sheet_nr)

    list_data = []
    index = 0
    names = []
    data = []
    for (names, data) in raw_data:
        print(data)
    #for (column_name, data) in raw_data:
     #   print("column: ", column_name , " data: ", data[0])
    #arrayofvalues = sheet.col_values(0, 1)
    #for row in (sheet.values):
    #    for value in row:
            #ws.cell(row=index, column=2).value = x1
            #if(value == "Kolumn1"):
    #        print(value)
    for col in sheet.iter_cols():
        for cell in col:
            print((cell.internal_value), "\n")
        #print("TEST: ", sheet[col[0]])
        #for row_index in range(1, sheet.nrows):
            #raw_data[col_index]
            #sheet.write(0, 0,'Inserting data in 1st Row and 1st Column')

    #wb_cpy = copy(workbook)
    workbook.save("joel_test.xlsx")
                

def start_setup():
        # raw_input returns the empty string for "enter"
        yes = set(['yes','y', 'ye', ''])
        no = set(['no','n'])
        exit = set(['exit!','quit!'])
        prompt = '> '
        #s"$file", "C:\\Users\\Looten\\Desktop\\python_test\\temp_excel.xlsx", "0", "doc created"

        print ("Hi, welcome to The best Excel parser known to man!")
        print ("Vilken fil vill du hämta data från?")
        #path_1 = "C:\\workspace\\python_test\\temp_excel.xlsx"#input(prompt)
        raw_file = "C:\\workspace\\python_test\\raw_data.xlsx"

        print ("Vilken fil är mallen?")
        plot_file = "C:\\workspace\\python_test\\plott_mall.xlsx"#input(prompt)

        print ("Vilket blad ligger data i? (Detta ska anges med nollindex! dvs Blad 1 är 0, Blad 2 är 1 etc")
        #sheet_nr = input(prompt)
        sheet_nr = "0"

        print ("Vad heter columnen?")
        #sheet_nr = "col1"
        col_name = "col1"
        #print ("What text or data do you want to exchange from the excel sheet?")
        #searched_text = input(prompt)
        #searched_text = "joel kan allt"

        print ("What text or data do you want it exchange to?")
        #exchange_text = input(prompt)

        #confirm = input("> Do you want to parse  " + path + "? (Yes/No)").lower()
        confirm = "yes"

        if confirm in yes:
            #parse_excel(path, int(sheet_nr), searched_text)
            raw_data = fetch_raw_data(raw_file, int(sheet_nr), col_name)
            insert_data_to_plot_file(plot_file, int(sheet_nr), raw_data)
            return True
        elif confirm in no:
            print ("Ok, nothing will happen !")
            return False
        else:
           print ("Please answer yes or no !, the setup will start again")
           # start again
           start_setup()

if __name__ == "__main__":
    start_setup()
    #print("iasda")
    #path = input("yeee")

#----------------------------------------------------------------
    # Detta används EJ
    # kan användas för att peka ut et specifikt blad i excel arket
    # worksheet = workbook.sheet_by_name('My_Sheet_Name')

    # idexerar bladen i excel arket
    # worksheet = workbook.sheet_by_index(0)

    #keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
    #dict_list = []
    #for row_index in range(1, sheet.nrows):
        #d = {keys[col_index]: sheet.cell(row_index, col_index).value for col_index in range(sheet.ncols)}
        #dict_list.append(d)
    #print(dict_list)

    #dict_list.append(d)
    #for index in workbook.nsheets:
        #sheet = workbook.sheet_by_index(index)
        #print("Blad:", sheet)
        # if sheet.cell(0, 0).value == xlrd.empty_cell.value:
        # hämtar värdet från en viss cell (rad, column)
        # value = sheet.cell(0, 0).value
        # sheet.write(0, 0,'Inserting data in 1st Row and 1st Column')
